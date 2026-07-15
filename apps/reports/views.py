"""
Reports app views: Employee submission, history, Admin management, exports.
"""
import csv
import io
import logging
from datetime import date

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.http import HttpResponse, Http404
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, View

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from apps.core.mixins import AdminRequiredMixin, EmployeeRequiredMixin, OwnershipMixin
from apps.core.utils import log_activity, get_client_ip, send_notification
from apps.employees.models import Department, EmployeeProfile
from django.db import models
from .models import DailyReport, Activity, DailyReportActivity
from .forms import DailyReportForm, AdminReportEditForm, ReportFilterForm, ActivityForm
from .utils import apply_date_filter

logger = logging.getLogger(__name__)
User = get_user_model()


# ── Employee: Submit Daily Report ─────────────────────────────────────────────

class ReportSubmitView(EmployeeRequiredMixin, CreateView):
    """Employees submit their daily report — one per day enforced."""

    model = DailyReport
    form_class = DailyReportForm
    template_name = 'reports/report_submit.html'
    success_url = reverse_lazy('dashboard:index')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        try:
            profile = request.user.employee_profile
        except EmployeeProfile.DoesNotExist:
            messages.error(request, 'No employee profile found. Contact your administrator.')
            return redirect('dashboard:index')

        today = date.today()
        if DailyReport.objects.filter(employee=profile, date=today).exists():
            messages.warning(request, 'You have already submitted today\'s report.')
            return redirect('reports:report_history')
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        profile = request.user.employee_profile
        today = date.today()

        if DailyReport.objects.filter(employee=profile, date=today).exists():
            messages.warning(request, 'You have already submitted today\'s report.')
            return redirect('reports:report_history')

        # Parse dynamic activity entries
        activities = request.POST.getlist('activity')
        reports = request.POST.getlist('report')
        quantities = request.POST.getlist('quantity')

        errors = []
        submitted_entries = []

        active_activities = Activity.objects.filter(is_active=True)
        active_ids = set(active_activities.values_list('id', flat=True))

        for i in range(len(activities)):
            act_id_str = activities[i].strip()
            rep_text = reports[i].strip()
            qty_str = quantities[i].strip()

            # Skip row if completely empty
            if not act_id_str and not rep_text and not qty_str:
                continue

            row_error = None
            if not act_id_str:
                row_error = "Please select an activity."
            else:
                try:
                    act_id = int(act_id_str)
                    if act_id not in active_ids:
                        row_error = "Selected activity is invalid or inactive."
                except ValueError:
                    row_error = "Invalid activity choice."

            if not rep_text:
                row_error = row_error or "Report details cannot be blank."
            elif len(rep_text) < 10:
                row_error = row_error or "Report details must be at least 10 characters."

            qty = None
            if not qty_str:
                row_error = row_error or "Quantity cannot be blank."
            else:
                try:
                    qty = int(qty_str)
                    if qty <= 0:
                        row_error = row_error or "Quantity must be a positive integer."
                except ValueError:
                    row_error = row_error or "Quantity must be a numeric value."

            if row_error:
                errors.append(f"Row {i+1}: {row_error}")

            submitted_entries.append({
                'activity_id': int(act_id_str) if act_id_str.isdigit() else '',
                'report_text': rep_text,
                'quantity': qty_str,
            })

        if not submitted_entries:
            errors.append("You must add at least one activity entry.")

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, self.template_name, {
                'form': self.get_form(),
                'active_activities': active_activities,
                'submitted_entries': submitted_entries,
            })

        import django.db.transaction as transaction
        try:
            with transaction.atomic():
                report = DailyReport.objects.create(
                    employee=profile,
                    date=today,
                    report_text="",
                )
                for entry in submitted_entries:
                    DailyReportActivity.objects.create(
                        daily_report=report,
                        activity_id=entry['activity_id'],
                        report_text=entry['report_text'],
                        quantity=int(entry['quantity']),
                    )
                report.save()

            log_activity(
                self.request.user, 'REPORT_CREATED',
                f"Report submitted for {report.date}.",
                get_client_ip(self.request),
            )
            for admin_user in User.objects.filter(is_staff=True):
                send_notification(
                    admin_user,
                    f"{profile.full_name} submitted their report for {report.date}.",
                    'info',
                    f'/reports/admin/{report.pk}/',
                )
            messages.success(self.request, 'Your daily report has been submitted successfully!')
            return redirect(self.success_url)
        except Exception as e:
            logger.exception("Error saving report")
            messages.error(request, f"An error occurred while saving your report: {e}")
            return render(request, self.template_name, {
                'form': self.get_form(),
                'active_activities': active_activities,
                'submitted_entries': submitted_entries,
            })

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['active_activities'] = Activity.objects.filter(is_active=True)
        ctx['submitted_entries'] = []
        return ctx



# ── Employee: Report History ──────────────────────────────────────────────────

class ReportHistoryView(EmployeeRequiredMixin, ListView):
    """Employees view only their own report history with filtering."""

    model = DailyReport
    template_name = 'reports/report_history.html'
    context_object_name = 'reports'
    paginate_by = 15

    def get_queryset(self):
        try:
            profile = self.request.user.employee_profile
        except EmployeeProfile.DoesNotExist:
            return DailyReport.objects.none()

        qs = DailyReport.objects.filter(employee=profile).order_by('-date')
        form = ReportFilterForm(self.request.GET)
        if form.is_valid():
            qs = apply_date_filter(
                qs,
                form.cleaned_data.get('date_range', ''),
                form.cleaned_data.get('date_from'),
                form.cleaned_data.get('date_to'),
            )
            keyword = form.cleaned_data.get('keyword', '')
            if keyword:
                qs = qs.filter(report_text__icontains=keyword)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = ReportFilterForm(self.request.GET or None)
        try:
            profile = self.request.user.employee_profile
            ctx['total_reports'] = DailyReport.objects.filter(employee=profile).count()
            ctx['this_month'] = DailyReport.objects.filter(
                employee=profile, date__year=date.today().year, date__month=date.today().month
            ).count()
        except EmployeeProfile.DoesNotExist:
            pass
        return ctx


class ReportDetailView(EmployeeRequiredMixin, OwnershipMixin, DetailView):
    """Employee views a specific one of their own reports."""

    model = DailyReport
    template_name = 'reports/report_detail.html'
    context_object_name = 'report'

    def get_queryset(self):
        return DailyReport.objects.select_related(
            'employee__user',
            'employee__department',
            'employee__designation',
        ).prefetch_related('activities__activity')


# ── Admin: Report Management ──────────────────────────────────────────────────

class AdminReportListView(AdminRequiredMixin, ListView):
    """Admin views all reports with full filtering."""

    model = DailyReport
    template_name = 'reports/admin_report_list.html'
    context_object_name = 'reports'
    paginate_by = 20

    def get_queryset(self):
        qs = DailyReport.objects.select_related('employee__user', 'employee__department').order_by('-date', '-created_at')
        form = ReportFilterForm(self.request.GET)
        if form.is_valid():
            qs = apply_date_filter(
                qs,
                form.cleaned_data.get('date_range', ''),
                form.cleaned_data.get('date_from'),
                form.cleaned_data.get('date_to'),
            )
            keyword = form.cleaned_data.get('keyword', '')
            if keyword:
                qs = qs.filter(report_text__icontains=keyword) | qs.filter(
                    employee__user__first_name__icontains=keyword
                ) | qs.filter(employee__user__last_name__icontains=keyword)
            dept = form.cleaned_data.get('department')
            if dept:
                qs = qs.filter(employee__department_id=dept)
            emp = form.cleaned_data.get('employee')
            if emp:
                qs = qs.filter(employee_id=emp)
        return qs.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['filter_form'] = ReportFilterForm(self.request.GET or None)
        ctx['departments'] = Department.objects.filter(is_active=True)
        ctx['employees'] = EmployeeProfile.objects.select_related('user').filter(is_active=True)
        ctx['total_count'] = self.get_queryset().count()
        # Pending employees (no report today)
        today = date.today()
        submitted_ids = DailyReport.objects.filter(date=today).values_list('employee_id', flat=True)
        ctx['pending_count'] = EmployeeProfile.objects.filter(is_active=True).exclude(id__in=submitted_ids).count()
        return ctx


class AdminReportDetailView(AdminRequiredMixin, DetailView):
    model = DailyReport
    template_name = 'reports/admin_report_detail.html'
    context_object_name = 'report'

    def get_queryset(self):
        return DailyReport.objects.select_related(
            'employee__user',
            'employee__department',
            'employee__designation',
        ).prefetch_related('activities__activity')


class AdminReportEditView(AdminRequiredMixin, UpdateView):
    model = DailyReport
    form_class = AdminReportEditForm
    template_name = 'reports/admin_report_edit.html'
    context_object_name = 'report'

    def get_success_url(self):
        return reverse_lazy('reports:admin_report_detail', kwargs={'pk': self.object.pk})

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()

        # Parse inputs
        activities = request.POST.getlist('activity')
        reports = request.POST.getlist('report')
        quantities = request.POST.getlist('quantity')

        errors = []
        submitted_entries = []
        active_activities = Activity.objects.filter(is_active=True)
        active_ids = set(active_activities.values_list('id', flat=True))

        for i in range(len(activities)):
            act_id_str = activities[i].strip()
            rep_text = reports[i].strip()
            qty_str = quantities[i].strip()

            if not act_id_str and not rep_text and not qty_str:
                continue

            row_error = None
            if not act_id_str:
                row_error = "Please select an activity."
            else:
                try:
                    act_id = int(act_id_str)
                    existing_activity_ids = set(self.object.activities.values_list('activity_id', flat=True))
                    if act_id not in active_ids and act_id not in existing_activity_ids:
                        row_error = "Selected activity is invalid."
                except ValueError:
                    row_error = "Invalid activity choice."

            if not rep_text:
                row_error = row_error or "Report details cannot be blank."
            elif len(rep_text) < 10:
                row_error = row_error or "Report details must be at least 10 characters."

            qty = None
            if not qty_str:
                row_error = row_error or "Quantity cannot be blank."
            else:
                try:
                    qty = int(qty_str)
                    if qty <= 0:
                        row_error = row_error or "Quantity must be a positive integer."
                except ValueError:
                    row_error = row_error or "Quantity must be a numeric value."

            if row_error:
                errors.append(f"Row {i+1}: {row_error}")

            submitted_entries.append({
                'activity_id': int(act_id_str) if act_id_str.isdigit() else '',
                'report_text': rep_text,
                'quantity': qty_str,
            })

        if not submitted_entries:
            errors.append("You must add at least one activity entry.")

        if not form.is_valid() or errors:
            for err in errors:
                messages.error(request, err)
            existing_act_ids = self.object.activities.values_list('activity_id', flat=True)
            all_selectable_activities = Activity.objects.filter(
                models.Q(is_active=True) | models.Q(id__in=existing_act_ids)
            ).distinct()
            return render(request, self.template_name, {
                'form': form,
                'object': self.object,
                'active_activities': all_selectable_activities,
                'submitted_entries': submitted_entries,
            })

        import django.db.transaction as transaction
        try:
            with transaction.atomic():
                report = form.save(commit=False)
                report.is_edited = True
                report.save()

                # Delete all current activity entries
                report.activities.all().delete()

                # Create new activity entries
                for entry in submitted_entries:
                    DailyReportActivity.objects.create(
                        daily_report=report,
                        activity_id=entry['activity_id'],
                        report_text=entry['report_text'],
                        quantity=int(entry['quantity']),
                    )
                report.save()

            log_activity(
                request.user, 'REPORT_UPDATED',
                f"Report #{report.pk} for {report.employee.full_name} updated.",
                get_client_ip(request),
            )
            send_notification(
                report.employee.user,
                f'Your report for {report.date} was edited by an administrator.',
                'warning',
                f'/reports/{report.pk}/',
            )
            messages.success(request, 'Report updated successfully.')
            return redirect(self.get_success_url())
        except Exception as e:
            logger.exception("Error updating report")
            messages.error(request, f"An error occurred while updating the report: {e}")
            existing_act_ids = self.object.activities.values_list('activity_id', flat=True)
            all_selectable_activities = Activity.objects.filter(
                models.Q(is_active=True) | models.Q(id__in=existing_act_ids)
            ).distinct()
            return render(request, self.template_name, {
                'form': form,
                'object': self.object,
                'active_activities': all_selectable_activities,
                'submitted_entries': submitted_entries,
            })

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        existing_act_ids = self.object.activities.values_list('activity_id', flat=True)
        ctx['active_activities'] = Activity.objects.filter(
            models.Q(is_active=True) | models.Q(id__in=existing_act_ids)
        ).distinct()
        ctx['submitted_entries'] = [
            {
                'activity_id': ra.activity_id,
                'report_text': ra.report_text,
                'quantity': ra.quantity,
            }
            for ra in self.object.activities.all()
        ]
        return ctx


class AdminReportDeleteView(AdminRequiredMixin, DeleteView):
    model = DailyReport
    template_name = 'reports/admin_report_confirm_delete.html'
    success_url = reverse_lazy('reports:admin_report_list')

    def form_valid(self, form):
        report = self.object
        log_activity(
            self.request.user, 'REPORT_DELETED',
            f"Report #{report.pk} for {report.employee.full_name} on {report.date} deleted.",
            get_client_ip(self.request),
        )
        resp = super().form_valid(form)
        messages.success(self.request, 'Report deleted successfully.')
        return resp


# ── Exports ───────────────────────────────────────────────────────────────────

def _get_filtered_reports(request):
    """Return filtered queryset for export, based on GET params."""
    qs = DailyReport.objects.select_related('employee__user', 'employee__department', 'employee__designation').order_by('-date')
    form = ReportFilterForm(request.GET)
    if form.is_valid():
        qs = apply_date_filter(qs, form.cleaned_data.get('date_range', ''), form.cleaned_data.get('date_from'), form.cleaned_data.get('date_to'))
        kw = form.cleaned_data.get('keyword', '')
        if kw:
            qs = qs.filter(report_text__icontains=kw)
        dept = form.cleaned_data.get('department')
        if dept:
            qs = qs.filter(employee__department_id=dept)
        emp = form.cleaned_data.get('employee')
        if emp:
            qs = qs.filter(employee_id=emp)
    return qs.distinct()


class ExportReportsCSVView(AdminRequiredMixin, View):
    """Export reports as CSV."""

    def get(self, request, *args, **kwargs):
        qs = _get_filtered_reports(request)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="reports.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Name', 'Department', 'Designation', 'Date', 'Words', 'Report', 'Submitted At'])
        for r in qs:
            writer.writerow([
                r.employee.employee_id,
                r.employee.full_name,
                r.employee.department.name if r.employee.department else '',
                r.employee.designation.name if r.employee.designation else '',
                r.date.strftime('%Y-%m-%d'),
                r.word_count,
                r.report_text,
                r.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
        log_activity(request.user, 'EXPORT', 'CSV export of reports.', get_client_ip(request))
        return response


class ExportReportsExcelView(AdminRequiredMixin, View):
    """Export reports as Excel (.xlsx)."""

    def get(self, request, *args, **kwargs):
        qs = _get_filtered_reports(request)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reports'

        headers = ['Employee ID', 'Name', 'Department', 'Designation', 'Date', 'Words', 'Report', 'Submitted At']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, r in enumerate(qs, 2):
            ws.cell(row=row_num, column=1, value=r.employee.employee_id)
            ws.cell(row=row_num, column=2, value=r.employee.full_name)
            ws.cell(row=row_num, column=3, value=r.employee.department.name if r.employee.department else '')
            ws.cell(row=row_num, column=4, value=r.employee.designation.name if r.employee.designation else '')
            ws.cell(row=row_num, column=5, value=r.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=r.word_count)
            ws.cell(row=row_num, column=7, value=r.report_text)
            ws.cell(row=row_num, column=8, value=r.created_at.strftime('%Y-%m-%d %H:%M:%S'))

        ws.column_dimensions['G'].width = 60
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'H']:
            ws.column_dimensions[col].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reports.xlsx"'
        log_activity(request.user, 'EXPORT', 'Excel export of reports.', get_client_ip(request))
        return response


class ExportReportsPDFView(AdminRequiredMixin, View):
    """Export reports as PDF."""

    def get(self, request, *args, **kwargs):
        qs = _get_filtered_reports(request)[:200]  # Limit PDF to 200 rows
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('Daily Reports Export', styles['Title']))
        elements.append(Spacer(1, 12))

        data = [['Employee', 'Department', 'Date', 'Words']]
        for r in qs:
            data.append([
                r.employee.full_name,
                r.employee.department.name if r.employee.department else 'N/A',
                r.date.strftime('%d %b %Y'),
                str(r.word_count),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reports.pdf"'
        log_activity(request.user, 'EXPORT', 'PDF export of reports.', get_client_ip(request))
        return response


# ── Admin: Activity Management ────────────────────────────────────────────────

class AdminActivityListView(AdminRequiredMixin, ListView):
    model = Activity
    template_name = 'reports/admin_activity_list.html'
    context_object_name = 'activities'
    paginate_by = 20

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(name__icontains=q)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search_query'] = self.request.GET.get('q', '')
        return ctx


class AdminActivityCreateView(AdminRequiredMixin, CreateView):
    model = Activity
    form_class = ActivityForm
    template_name = 'reports/admin_activity_form.html'
    success_url = reverse_lazy('reports:admin_activity_list')

    def form_valid(self, form):
        resp = super().form_valid(form)
        log_activity(self.request.user, 'ACTIVITY_CREATED', f"Activity '{self.object.name}' created.", get_client_ip(self.request))
        messages.success(self.request, f"Activity '{self.object.name}' created successfully.")
        return resp


class AdminActivityUpdateView(AdminRequiredMixin, UpdateView):
    model = Activity
    form_class = ActivityForm
    template_name = 'reports/admin_activity_form.html'
    success_url = reverse_lazy('reports:admin_activity_list')

    def form_valid(self, form):
        resp = super().form_valid(form)
        log_activity(self.request.user, 'ACTIVITY_UPDATED', f"Activity '{self.object.name}' updated.", get_client_ip(self.request))
        messages.success(self.request, f"Activity '{self.object.name}' updated successfully.")
        return resp


class AdminActivityDeleteView(AdminRequiredMixin, DeleteView):
    model = Activity
    template_name = 'reports/admin_activity_confirm_delete.html'
    success_url = reverse_lazy('reports:admin_activity_list')

    def form_valid(self, form):
        name = self.object.name
        resp = super().form_valid(form)
        log_activity(self.request.user, 'ACTIVITY_DELETED', f"Activity '{name}' deleted.", get_client_ip(self.request))
        messages.success(self.request, f"Activity '{name}' deleted successfully.")
        return resp

